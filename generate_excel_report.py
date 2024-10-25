import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.formatting.rule import CellIsRule

RESULTS_FILE = "test_results.log"
ERROR_LOGS_DIR = "error_logs"
OUTPUT_EXCEL = "test_report.xlsx"

TIME_REGEX = re.compile(r'(\d+[\.,]?\d*)s')

def parse_time(time_str):
    """
    Parses a time string formatted as 'X.XXXs' or 'X,XXXs' and returns a float.
    Replaces commas with dots to ensure correct parsing.
    """
    if not isinstance(time_str, str):
        return 0.0
    time_str = time_str.replace(',', '.')
    match = TIME_REGEX.search(time_str)
    return float(match.group(1)) if match else 0.0

def clean_excel_string(s):
    """
    Removes or replaces illegal characters that cannot be used in Excel worksheets.
    Allowed characters are typically Unicode characters above 31, excluding control characters.
    """
    if not isinstance(s, str):
        return s
    return ''.join(c if c >= ' ' or c in '\t\n\r' else ' ' for c in s)

def read_test_results(results_file):
    """
    Reads the test results from a CSV file and returns two DataFrames:
    1. test_cases_df: DataFrame containing individual test case results.
    2. suite_summary_df: DataFrame containing suite summary results.
    Handles missing or malformed fields by filling them with default values.
    """
    expected_columns = ['Test Suite', 'Test Case', 'Status', 'Time', 'Warnings', 'Errors', 'Skipped']

    try:
        print(f"Attempting to read '{results_file}' with delimiter: ','")
        df = pd.read_csv(
            results_file,
            sep=',',                        
            header=0,                       
            usecols=expected_columns,       
            skip_blank_lines=True
        )
        print("Successfully read the file with comma delimiter.")
    except pd.errors.ParserError as e:
        print(f"ParserError: {e}")
        return pd.DataFrame(), pd.DataFrame()
    except Exception as e:
        print(f"Unexpected error while reading the file: {e}")
        return pd.DataFrame(), pd.DataFrame()

    print("First 5 rows of the parsed DataFrame:")
    print(df.head())

    suite_summary_df = df[df['Test Case'].str.strip().str.lower() == 'suite summary']
    test_cases_df = df[df['Test Case'].str.strip().str.lower() != 'suite summary']

    print(f"Number of Suite Summaries: {len(suite_summary_df)}")
    print("Suite Summaries DataFrame:")
    print(suite_summary_df.head())

    print(f"Number of Test Cases: {len(test_cases_df)}")
    print("Test Cases DataFrame:")
    print(test_cases_df.head())

    malformed_test_cases = test_cases_df[test_cases_df.isnull().any(axis=1)]
    if not malformed_test_cases.empty:
        print("Warning: Found malformed Test Case rows with missing fields. These rows will have missing values filled with defaults:")
        print(malformed_test_cases)
        test_cases_df['Time'] = test_cases_df['Time'].fillna("0s")
        test_cases_df['Warnings'] = test_cases_df['Warnings'].fillna(0)
        test_cases_df['Errors'] = test_cases_df['Errors'].fillna(0)
        test_cases_df['Skipped'] = test_cases_df['Skipped'].fillna(0)

    malformed_suite_summaries = suite_summary_df[suite_summary_df.isnull().any(axis=1)]
    if not malformed_suite_summaries.empty:
        print("Warning: Found malformed Suite Summary rows with missing fields. These rows will have missing values filled with defaults:")
        print(malformed_suite_summaries)
        suite_summary_df['Time'] = suite_summary_df['Time'].fillna("0s")
        suite_summary_df['Warnings'] = suite_summary_df['Warnings'].fillna(0)
        suite_summary_df['Errors'] = suite_summary_df['Errors'].fillna(0)
        suite_summary_df['Skipped'] = suite_summary_df['Skipped'].fillna(0)

    test_cases_df['Time (s)'] = test_cases_df['Time'].apply(parse_time)
    suite_summary_df['Time (s)'] = suite_summary_df['Time'].apply(parse_time)

    numeric_fields = ['Warnings', 'Errors', 'Skipped']
    for field in numeric_fields:
        test_cases_df[field] = pd.to_numeric(test_cases_df[field], errors='coerce').fillna(0).astype(int)
        suite_summary_df[field] = pd.to_numeric(suite_summary_df[field], errors='coerce').fillna(0).astype(int)

    return test_cases_df, suite_summary_df

def generate_summary(test_cases_df, suite_summary_df):
    """
    Generates a summary dictionary containing counts of passed, skipped, and failed tests.
    Also aggregates warnings, errors, and time from suite summaries.
    """
    status_priority = {'ERROR':4, 'FAILED':3, 'FAIL':3, 'SKIPPED':2, 'PASSED':1, 'PASS':1}

    test_cases_df = test_cases_df.copy()
    test_cases_df['Status_Priority'] = test_cases_df['Status'].str.upper().map(status_priority).fillna(0)

    final_status = test_cases_df.groupby('Test Case').agg({'Status_Priority':'max', 'Status':'first'}).reset_index()

    def resolve_status(row):
        if row['Status_Priority'] >=4:
            return 'ERROR'
        elif row['Status_Priority'] >=3:
            return 'FAILED'
        elif row['Status_Priority'] >=2:
            return 'SKIPPED'
        elif row['Status_Priority'] >=1:
            return 'PASSED'
        else:
            return 'UNKNOWN'

    final_status['Final Status'] = final_status.apply(resolve_status, axis=1)

    passed = final_status[final_status['Final Status'] == 'PASSED']['Test Case'].nunique()
    skipped = final_status[final_status['Final Status'] == 'SKIPPED']['Test Case'].nunique()
    failed_test_cases = final_status[final_status['Final Status'].isin(['FAILED', 'ERROR'])]['Test Case'].nunique()

    total_warnings = suite_summary_df['Warnings'].sum()
    total_errors = suite_summary_df['Errors'].sum()
    total_time = suite_summary_df['Time (s)'].sum()

    total_test_suites = suite_summary_df['Test Suite'].nunique()

    total_failed = failed_test_cases + total_errors

    summary = {
        'Total Test Suites': total_test_suites,
        'Total Test Cases': final_status['Test Case'].nunique(),
        'Total Passed': passed,
        'Total Skipped': skipped,
        'Total Failed': total_failed,
        'Total Warnings': total_warnings,
        'Total Errors': total_errors,
        'Total Time (s)': total_time
    }

    print(f"Passed: {passed}, Skipped: {skipped}, Failed Test Cases: {failed_test_cases}, Total Test Cases: {final_status['Test Case'].nunique()}")
    print(f"Total Warnings: {total_warnings}, Total Errors: {total_errors}, Total Time (s): {total_time}")
    print(f"Total Failed (Including Suite Errors): {total_failed}")

    return summary

def create_excel_report(test_cases_df, suite_summary_df, summary, output_file):
    """
    Creates an Excel report with Summary, Detailed Results, Suite Summaries, and Error Logs sheets.
    """
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        summary_df = pd.DataFrame(list(summary.items()), columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        detailed_df = test_cases_df.copy()
        detailed_df['Status'] = detailed_df['Status'].str.upper()
        if 'Status_Priority' in detailed_df.columns:
            detailed_df = detailed_df.drop(columns=['Status_Priority'])
        detailed_df.to_excel(writer, sheet_name='Detailed Results', index=False)

        suite_summary_df.to_excel(writer, sheet_name='Suite Summaries', index=False)

        error_tests = pd.concat([
            test_cases_df[test_cases_df['Status'].str.upper().isin(['FAILED', 'ERROR'])],
            suite_summary_df[suite_summary_df['Status'].str.upper().isin(['FAILED', 'ERROR'])]
        ])

        print(f"Number of error tests (Test Cases + Suite Summaries): {len(error_tests)}")

        if not error_tests.empty:
            error_logs = []
            for _, row in error_tests.iterrows():
                test_file = row['Test Suite']
                test_case = row['Test Case']
                status = row['Status']
                test_name = os.path.basename(test_file).replace('.py', '')
                error_log_file = f"{test_name}_error.log"
                error_log_path = os.path.join(ERROR_LOGS_DIR, error_log_file)

                print(f"Looking for error log: {error_log_path}")

                if os.path.exists(error_log_path):
                    with open(error_log_path, 'r') as f:
                        error_content = f.read()
                    print(f"Found error log: {error_log_file}")
                else:
                    error_content = "Error log file not found."
                    print(f"Error log not found: {error_log_file}")

                error_content_clean = clean_excel_string(error_content)
                error_logs.append({
                    'Test Suite': test_file,
                    'Test Case': test_case,
                    'Status': status,
                    'Error Details': error_content_clean
                })
            error_logs_df = pd.DataFrame(error_logs)
            error_logs_df.to_excel(writer, sheet_name='Error Logs', index=False)
        else:
            print("No errors found to include in the Error Logs sheet.")

    apply_formatting(output_file, summary)

def apply_formatting(excel_file, summary):
    """
    Applies formatting to the Excel report for better readability and visualization.
    """
    wb = load_workbook(excel_file)

    summary_sheet = wb['Summary']
    for cell in summary_sheet["1:1"]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for column in summary_sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2)
        summary_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    pie = PieChart()
    labels = Reference(summary_sheet, min_col=1, min_row=3, max_row=5)
    data = Reference(summary_sheet, min_col=2, min_row=3, max_row=5)
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    pie.title = "Test Status Distribution"
    summary_sheet.add_chart(pie, "D2")

    detailed_sheet = wb['Detailed Results']
    for cell in detailed_sheet["1:1"]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for column in detailed_sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2)
        detailed_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    status_fills = {
        'PASSED': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        'FAILED': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'ERROR': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'SKIPPED': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    }
    for row in detailed_sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            status = cell.value
            if pd.isna(status):
                continue
            fill = status_fills.get(status.upper(), None)
            if fill:
                cell.fill = fill
    for row in detailed_sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")
    red_fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
    time_col_letter = None
    for cell in detailed_sheet[1]:
        if cell.value == 'Time (s)':
            time_col_letter = get_column_letter(cell.column)
            break
    if time_col_letter:
        range_str = f"{time_col_letter}2:{time_col_letter}{detailed_sheet.max_row}"
        detailed_sheet.conditional_formatting.add(
            range_str,
            CellIsRule(operator='greaterThan', formula=['1.0'], fill=red_fill)
        )

    suite_summaries_sheet = wb['Suite Summaries']
    for cell in suite_summaries_sheet["1:1"]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    for column in suite_summaries_sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2)
        suite_summaries_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    status_fills_suite = {
        'FAILED': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'ERROR': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
        'SKIPPED': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        'PASS': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    }
    for row in suite_summaries_sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            status = cell.value
            if pd.isna(status):
                continue
            fill = status_fills_suite.get(status.upper(), None)
            if fill:
                cell.fill = fill
    for row in suite_summaries_sheet.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    if 'Error Logs' in wb.sheetnames:
        error_sheet = wb['Error Logs']
        for cell in error_sheet["1:1"]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        for column in error_sheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2)
            error_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        for row in error_sheet.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
    wb.save(excel_file)
    print(f"Excel report generated successfully: {excel_file}")

def main():
    if not os.path.exists(RESULTS_FILE):
        print(f"Results file not found: {RESULTS_FILE}")
        return

    test_cases_df, suite_summary_df = read_test_results(RESULTS_FILE)

    if test_cases_df.empty and suite_summary_df.empty:
        print("No valid test results found to generate the report.")
        return

    summary = generate_summary(test_cases_df, suite_summary_df)
    create_excel_report(test_cases_df, suite_summary_df, summary, OUTPUT_EXCEL)

if __name__ == "__main__":
    main()
