import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.marker import DataPoint

RESULTS_FILE = "test_results.log"
ERROR_LOGS_DIR = "error_logs"
OUTPUT_EXCEL = "test_report.xlsx"

TIME_REGEX = re.compile(r'(\d+[\.,]?\d*)s')

def parse_time(time_str):
    if not isinstance(time_str, str):
        return 0.0
    time_str = time_str.replace(',', '.')
    match = TIME_REGEX.search(time_str)
    return float(match.group(1)) if match else 0.0

def clean_excel_string(s):
    if not isinstance(s, str):
        return s
    return ''.join(c if c >= ' ' or c in '\t\n\r' else ' ' for c in s)

def read_test_results(results_file):
    expected_columns = ['Test Suite', 'Test Case', 'Status', 'Time', 'Warnings', 'Errors', 'Skipped']
    try:
        df = pd.read_csv(results_file, sep=',', header=0, usecols=expected_columns, skip_blank_lines=True)
    except Exception as e:
        return pd.DataFrame(), pd.DataFrame()

    suite_summary_df = df[df['Test Case'].str.strip().str.lower() == 'suite summary'].copy()
    test_cases_df = df[df['Test Case'].str.strip().str.lower() != 'suite summary'].copy()

    test_cases_df.loc[:, 'Time'] = test_cases_df['Time'].fillna("0s")
    suite_summary_df.loc[:, 'Time'] = suite_summary_df['Time'].fillna("0s")
    
    test_cases_df.loc[:, 'Time (s)'] = test_cases_df['Time'].apply(parse_time)
    suite_summary_df.loc[:, 'Time (s)'] = suite_summary_df['Time'].apply(parse_time)

    numeric_fields = ['Warnings', 'Errors', 'Skipped']
    for field in numeric_fields:
        test_cases_df.loc[:, field] = pd.to_numeric(test_cases_df[field], errors='coerce').fillna(0).astype(int)
        suite_summary_df.loc[:, field] = pd.to_numeric(suite_summary_df[field], errors='coerce').fillna(0).astype(int)

    return test_cases_df, suite_summary_df

def generate_summary(test_cases_df, suite_summary_df):
    status_priority = {'ERROR': 4, 'FAILED': 3, 'FAIL': 3, 'SKIPPED': 2, 'PASSED': 1, 'PASS': 1}
    test_cases_df['Status_Priority'] = test_cases_df['Status'].str.upper().map(status_priority).fillna(0)
    final_status = test_cases_df.groupby('Test Case').agg({'Status_Priority': 'max', 'Status': 'first'}).reset_index()

    def resolve_status(row):
        if row['Status_Priority'] >= 4:
            return 'FAILED'
        elif row['Status_Priority'] >= 3:
            return 'FAILED'
        elif row['Status_Priority'] >= 2:
            return 'SKIPPED'
        elif row['Status_Priority'] >= 1:
            return 'PASSED'
        else:
            return 'UNKNOWN'

    final_status['Final Status'] = final_status.apply(resolve_status, axis=1)
    passed_test_cases = final_status[final_status['Final Status'] == 'PASSED']['Test Case'].nunique()
    skipped_test_cases = final_status[final_status['Final Status'] == 'SKIPPED']['Test Case'].nunique()
    failed_test_cases = final_status[final_status['Final Status'] == 'FAILED']['Test Case'].nunique()

    suite_summary_df['Status'] = suite_summary_df['Status'].str.upper()
    passed_suites = suite_summary_df[suite_summary_df['Status'] == 'PASS']['Test Suite'].nunique()
    skipped_suites = suite_summary_df[suite_summary_df['Status'] == 'SKIPPED']['Test Suite'].nunique()
    failed_suites = suite_summary_df[suite_summary_df['Status'].isin(['ERROR', 'FAIL'])]['Test Suite'].nunique()

    total_warnings = suite_summary_df['Warnings'].sum()
    total_errors = suite_summary_df['Errors'].sum()
    total_time = suite_summary_df['Time (s)'].sum()
    total_test_suites = suite_summary_df['Test Suite'].nunique()

    total_failed = failed_test_cases + failed_suites
    total_skipped = skipped_test_cases + skipped_suites
    total_passed = passed_test_cases + passed_suites
    total_test_cases = final_status['Test Case'].nunique()

    summary = {
        'Test Suites': {
            'Total Test Suites': total_test_suites,
            'Passed': passed_suites,
            'Skipped': skipped_suites,
            'Failed': failed_suites
        },
        'Test Cases': {
            'Total Test Cases': total_test_cases,
            'Passed': passed_test_cases,
            'Skipped': skipped_test_cases,
            'Failed': failed_test_cases
        },
        'Overall Summary': {
            'Total Passed': total_passed,
            'Total Skipped': total_skipped,
            'Total Failed': total_failed,
            'Total Warnings': total_warnings,
            'Total Errors': total_errors,
            'Total Time (s)': total_time
        }
    }

    return summary

def create_excel_report(test_cases_df, suite_summary_df, summary, output_file):
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        suites_summary_df = pd.DataFrame(list(summary['Test Suites'].items()), columns=['Metric', 'Value'])
        test_cases_summary_df = pd.DataFrame(list(summary['Test Cases'].items()), columns=['Metric', 'Value'])
        overall_summary_df = pd.DataFrame(list(summary['Overall Summary'].items()), columns=['Metric', 'Value'])

        suites_summary_df.to_excel(writer, sheet_name='Summary', index=False, startrow=0)
        test_cases_summary_df.to_excel(writer, sheet_name='Summary', index=False, startrow=5)
        overall_summary_df.to_excel(writer, sheet_name='Summary', index=False, startrow=10)

        detailed_df = test_cases_df.copy()
        detailed_df['Status'] = detailed_df['Status'].str.upper()
        detailed_df.to_excel(writer, sheet_name='Detailed Results', index=False)
        suite_summary_df.to_excel(writer, sheet_name='Suite Summaries', index=False)

        error_tests = pd.concat([
            test_cases_df[test_cases_df['Status'].str.upper().isin(['FAILED', 'ERROR'])],
            suite_summary_df[suite_summary_df['Status'].str.upper().isin(['ERROR', 'FAIL'])]
        ])

        if not error_tests.empty:
            error_logs = []
            for _, row in error_tests.iterrows():
                test_file = row['Test Suite']
                test_case = row['Test Case']
                status = row['Status']
                test_name = os.path.basename(test_file).replace('.py', '')
                error_log_file = f"{test_name}_error.log"
                error_log_path = os.path.join(ERROR_LOGS_DIR, error_log_file)

                if os.path.exists(error_log_path):
                    with open(error_log_path, 'r') as f:
                        error_content = f.read()
                else:
                    error_content = "Error log file not found."

                # Clean the error content and treat it as plain text
                error_content_clean = clean_excel_string(error_content)
                error_logs.append({
                    'Test Suite': test_file,
                    'Test Case': test_case,
                    'Status': status,
                    'Error Details': error_content_clean
                })
            error_logs_df = pd.DataFrame(error_logs)
            error_logs_df.to_excel(writer, sheet_name='Error Logs', index=False)

    apply_formatting(output_file, summary)

def apply_formatting(excel_file, summary):
    wb = load_workbook(excel_file)

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    alignment_center = Alignment(horizontal="center")

    summary_sheet = wb['Summary']
    for cell in summary_sheet["A1:B1"]:
        for c in cell:
            c.font = header_font
            c.fill = header_fill
            c.alignment = alignment_center
    for cell in summary_sheet["A6:B6"]:
        for c in cell:
            c.font = header_font
            c.fill = header_fill
            c.alignment = alignment_center
    for cell in summary_sheet["A11:B11"]:
        for c in cell:
            c.font = header_font
            c.fill = header_fill
            c.alignment = alignment_center

    summary_sheet.column_dimensions["A"].width = 20
    summary_sheet.column_dimensions["B"].width = 40

    pie_suites = PieChart()
    labels_suites = Reference(summary_sheet, min_col=1, min_row=3, max_row=5)
    data_suites = Reference(summary_sheet, min_col=2, min_row=3, max_row=5)
    pie_suites.add_data(data_suites, titles_from_data=False)
    pie_suites.set_categories(labels_suites)
    pie_suites.title = "Test Suites Status Distribution"

    pie_cases = PieChart()
    labels_cases = Reference(summary_sheet, min_col=1, min_row=8, max_row=10)
    data_cases = Reference(summary_sheet, min_col=2, min_row=8, max_row=10)
    pie_cases.add_data(data_cases, titles_from_data=False)
    pie_cases.set_categories(labels_cases)
    pie_cases.title = "Test Cases Status Distribution"

    colors = ["A8D5BA", "FFE8A1", "F7AFAF"]
    for i, color in enumerate(colors):
        dp_suites = DataPoint(idx=i)
        dp_suites.graphicalProperties.solidFill = color
        pie_suites.series[0].dPt.append(dp_suites)

        dp_cases = DataPoint(idx=i)
        dp_cases.graphicalProperties.solidFill = color
        pie_cases.series[0].dPt.append(dp_cases)

    summary_sheet.add_chart(pie_suites, "E1")
    summary_sheet.add_chart(pie_cases, "E20")

    def style_sheet(ws, header_row=1):
        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment_center

        status_colors = {
            'PASS': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'PASSED': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
            'SKIPPED': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
            'FAIL': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
            'ERROR': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        }
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                fill = status_colors.get(cell.value.upper(), None) if isinstance(cell.value, str) else None
                if fill:
                    cell.fill = fill
                cell.alignment = alignment_center

    style_sheet(summary_sheet)
    style_sheet(wb['Detailed Results'])
    style_sheet(wb['Suite Summaries'])
    style_sheet(wb['Error Logs'])

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for column in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
            adjusted_width = max_length + 15
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    total_passed = f"{summary['Test Cases']['Passed']} Cases from {summary['Test Suites']['Passed']} Suites"
    total_skipped = f"{summary['Test Suites']['Skipped']} Suites and {summary['Test Cases']['Skipped']} Cases"
    total_failed = f"{summary['Test Suites']['Failed']} Suites and {summary['Test Cases']['Failed']} Cases"

    summary_sheet.cell(row=12, column=2, value=total_passed)
    summary_sheet.cell(row=13, column=2, value=total_skipped)
    summary_sheet.cell(row=14, column=2, value=total_failed)

    wb.save(excel_file)

def main():
    if not os.path.exists(RESULTS_FILE):
        print(f"Results file not found: {RESULTS_FILE}")
        return

    test_cases_df, suite_summary_df = read_test_results(RESULTS_FILE)
    if test_cases_df.empty and suite_summary_df.empty:
        print("No valid test results found to generate the report.")
        return

    summary = generate_summary(test_cases_df, suite_summary_df)
    create_excel_report(test_cases_df, suite_summary_df, summary, OUTPUT_EXCEL)

if __name__ == "__main__":
    main()
